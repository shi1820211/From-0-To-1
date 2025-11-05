import json
import itertools
from typing import Dict, List, Union, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class RuleEngine:
    def __init__(self, excel_path: str):
        self.admission_rules = {}  # 准入规则列A
        self.rule_units = {}  # 准入规则单位列B
        self.input_units = {}  # 输入单位列E
        self.gender_conditions = {}  # 适用性别列I
        self.age_conditions = {}  # 适用年龄列J
        self.physiological_conditions = {}  # 生理状态条件
        self.disease_conditions = {}  # 疾病条件
        self.complex_rules = {}  # 规则列G
        self.excel_path = excel_path  # Excel文件路径

    def set_admission_rules(self, rules: Dict[str, str]):
        self.admission_rules = rules

    def set_rule_units(self, units: Dict[str, Dict[str, str]]):
        self.rule_units = units

    def set_input_units(self, units: Dict[str, str]):
        self.input_units = units

    def set_gender_conditions(self, conditions: Dict[str, str]):
        self.gender_conditions = conditions

    def set_age_conditions(self, conditions: Dict[str, str]):
        self.age_conditions = conditions

    def set_physiological_conditions(self, conditions: Dict[str, List[str]]):
        """设置生理状态条件，值来自准入规则A列"""
        self.physiological_conditions = {
            rule_id: [value for value in values if value in self.admission_rules.values()]
            for rule_id, values in conditions.items()
        }

    def set_disease_conditions(self, conditions: Dict[str, List[str]]):
        """设置疾病条件，值来自准入规则A列"""
        self.disease_conditions = {
            rule_id: [value for value in values if value in self.admission_rules.values()]
            for rule_id, values in conditions.items()
        }

    def set_complex_rules(self, rules: Dict[str, str]):
        self.complex_rules = rules

    def _parse_condition(self, condition: str) -> Dict[str, Union[str, Dict]]:
        parts = condition.strip().split()
        if len(parts) != 3:
            raise ValueError(f"Invalid condition format: {condition}")

        uuid, operator, value = parts
        return {
            "uuid": uuid,
            "operator": operator,
            "value": value.strip("'\"") if value.startswith(("'", '"')) else value,
            "raw_condition": condition
        }

    def _build_expression(self, rule: str) -> Dict[str, List[List[Dict]]]:
        and_conditions = []
        condition_explanations = []

        for or_group in rule.split(' and '):
            or_conditions = []
            explanations = []
            for condition in or_group.split(' or '):
                condition = condition.strip('() ')
                if condition:
                    parsed = self._parse_condition(condition)
                    or_conditions.append(parsed)
                    explanations.append(f"{parsed['uuid']} {parsed['operator']} {parsed['valu e']}")
            if or_conditions:
                and_conditions.append(or_conditions)
                condition_explanations.append(" OR ".join(explanations))

        return {
            "conditions": and_conditions,
            "explanation": " AND ".join(condition_explanations)
        }

    def _generate_test_cases(self, and_conditions: List[List[Dict]]) -> List[Dict]:
        test_cases = []

        # 生成符合条件(true)的用例
        for combination in itertools.product(*and_conditions):
            test_case = {
                "items": {},
                "case_type": "true"
            }
            for condition in combination:
                test_case["items"][condition['uuid']] = {
                    "name": "",
                    "values": condition['value'],
                    "qualitativeResult": condition['value'],
                    "units": [],
                    "interval": "1-1"
                }
            test_cases.append(test_case)

        # 生成不符合条件(false)的用例
        for i, or_group in enumerate(and_conditions):
            for condition in or_group:
                false_case = {
                    "items": {},
                    "case_type": "false"
                }

                for j, other_group in enumerate(and_conditions):
                    if i != j:
                        other_condition = other_group[0]
                        false_case["items"][other_condition['uuid']] = {
                            "name": "",
                            "values": other_condition['value'],
                            "qualitativeResult": other_condition['value'],
                            "units": [],
                            "interval": "1-1"
                        }

                false_value = "不符合_" + condition['value']
                false_case["items"][condition['uuid']] = {
                    "name": "",
                    "values": false_value,
                    "qualitativeResult": false_value,
                    "units": [],
                    "interval": "1-1"
                }

                test_cases.append(false_case)

        return test_cases

    def _apply_gender_condition(self, test_case: Dict, rule_id: str) -> Dict:
        if rule_id in self.gender_conditions:
            test_case['pat_sex'] = self.gender_conditions[rule_id]
        return test_case

    def _apply_age_condition(self, test_case: Dict, rule_id: str) -> Dict:
        if rule_id in self.age_conditions:
            age_expr = self.age_conditions[rule_id]
            if '%s' in age_expr:
                test_case['age'] = {"value": age_expr.replace('%s', '')}
        return test_case

    def _apply_physiological_condition(self, test_case: Dict, rule_id: str) -> Dict:
        if rule_id in self.physiological_conditions:
            test_case['physiologicalState'] = [
                value for value in self.physiological_conditions[rule_id]
                if value in self.admission_rules.values()
            ]
        return test_case

    def _apply_disease_condition(self, test_case: Dict, rule_id: str) -> Dict:
        if rule_id in self.disease_conditions:
            test_case['disease'] = [
                value for value in self.disease_conditions[rule_id]
                if value in self.admission_rules.values()
            ]
        return test_case

    def generate_standard_input(self, rule_id: str) -> List[Dict]:
        if rule_id not in self.complex_rules:
            return []

        rule = self.complex_rules[rule_id]
        expression = self._build_expression(rule)
        test_cases = self._generate_test_cases(expression["conditions"])

        standard_inputs = []
        for case in test_cases:
            standard_input = {
                "input": json.dumps({
                    "items": case["items"],
                    "pat_sex": "",
                    "age": {},
                    "physiologicalState": [],
                    "disease": []
                }, ensure_ascii=False),
                "description": expression["explanation"],
                "case_type": case["case_type"]
            }
            standard_input = self._apply_gender_condition(standard_input, rule_id)
            standard_input = self._apply_age_condition(standard_input, rule_id)
            standard_input = self._apply_physiological_condition(standard_input, rule_id)
            standard_input = self._apply_disease_condition(standard_input, rule_id)
            standard_inputs.append(standard_input)

        return standard_inputs

    def save_to_excel(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "TestCases"

        headers = ["规则ID", "规则说明", "用例类型", "标准输入"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        row_num = 2
        for rule_id in self.complex_rules:
            standard_inputs = self.generate_standard_input(rule_id)
            for input_data in standard_inputs:
                ws.cell(row=row_num, column=1, value=rule_id)
                ws.cell(row=row_num, column=2, value=input_data['description'])
                ws.cell(row=row_num, column=3, value=input_data['case_type'])
                ws.cell(row=row_num, column=4, value=input_data['input'])
                row_num += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(self.excel_path)


if __name__ == "__main__":
    excel_path = "test_cases.xlsx"
    engine = RuleEngine(excel_path)

    # 设置准入规则A列的值
    engine.set_admission_rules({
        "rule1": "08dbbe57-a59a-4dbb-8db0-5a25caa9224d",
        "rule2": "N28.901",
        "rule3": "pregnant",
        "rule4": "diabetes"
    })

    engine.set_complex_rules({
        "rule1": "3cde5936-8a03-460e-87a7-02bcd42c5b73 == '升高' and (97f87e32-09a8-4ea5-9f57-1ff2209a3f2a == '降低' or 5afbd8f0-4e23-4dea-9ebc-8c094ac13211 == '降低')",
        "rule2": "5745fc33-4b75-4a34-ac58-8d04b2291306 == '阳性' and 08dae3f4-642f-4276-8e7e-05416daa9459 == '阳性'"
    })

    engine.set_gender_conditions({
        "rule1": "female",
        "rule2": "male"
    })

    engine.set_age_conditions({
        "rule1": "7<=%s<=16",
        "rule2": "%s>=18"
    })

    # 设置生理状态和疾病条件，值来自准入规则A列
    engine.set_physiological_conditions({
        "rule1": ["08dbbe57-a59a-4dbb-8db0-5a25caa9224d", "pregnant"],
        "rule2": ["N28.901"]
    })

    engine.set_disease_conditions({
        "rule1": ["diabetes"],
        "rule2": ["N28.901"]
    })

    engine.save_to_excel()