import openpyxl
import json
import random
from collections import OrderedDict
import pandas as pd

def generate_json_into_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        col_a, col_b, col_g = row[0], row[1], row[6]

        # -------------------------
        # Step 1: 处理 items
        # -------------------------
        items_dict = OrderedDict()
        if col_g:
            # 先分割表达式
            expressions = str(col_g).replace(" and ", "||").replace(" or ", "||").split("||")
            for expr in expressions:
                expr = expr.strip()
                if not expr:
                    continue
                parts = expr.split()
                if len(parts) >= 3:
                    uuid = parts[0].strip("()")
                    operator = parts[1].strip()
                    raw_value = parts[2].strip("\",(),\'")  # 去掉引号和括号
                    value = raw_value


                    # 数值判断
                    try:
                        num_val = float(raw_value)
                        if operator == ">":
                            value = num_val + 0.1
                        elif operator == "<":
                            value = num_val - 0.1
                        else:
                            value = num_val
                    except ValueError:
                        # 如果不是数值（例如“升高”“降低”），就保持原样
                        value = raw_value

                    # 构造 item
                    items_dict[uuid] = {
                        "name": "",
                        "values": value,
                        "qualitativeResult": value,
                        "units": [],
                        "interval": "1-1"
                    }

        # -------------------------
        # Step 2: 处理性别 (pat_sex)
        # -------------------------
        pat_sex = None
        if col_a:
            if "sex" in str(col_a):
                if "female" in str(col_a):
                    pat_sex = "female"
                elif "male" in str(col_a):
                    pat_sex = "male"

        # 如果没有性别限制，默认 female
        if not pat_sex:
            pat_sex = "female"

        # -------------------------
        # Step 3: 处理 physiologicalState
        # -------------------------
        physiological_state = []
        if col_a and "in physiologicalState" in str(col_a):
            try:
                uuid = str(col_a).split("'")[1]
                physiological_state.append(uuid)
            except IndexError:
                pass

        # -------------------------
        # Step 4: 处理 age
        # -------------------------
        age = {}
        if col_b:  # 如果 B 列有单位
            unit_dict = eval(col_b) if isinstance(col_b, str) else col_b
            for unit in unit_dict.values():
                if unit == "month":
                    # 新逻辑：0–1 之间随机小数，保留 1 位
                    age[unit] = round(random.uniform(0, 1), 1)
                else:
                    # 默认整数 0–150
                    age[unit] = random.randint(0, 150)
        else:
            # B 列为空，默认 year
            age["year"] = random.randint(0, 150)

        # -------------------------
        # Step 5: disease
        # -------------------------
        disease = []

        # -------------------------
        # 组合结果
        # -------------------------
        result_obj = {
            "input": json.dumps({
                "items": items_dict,
                "pat_sex": pat_sex,
                "age": age,
                "physiologicalState": physiological_state,
                "disease": disease
            }, ensure_ascii=False)
        }

        results.append(result_obj)

    # -------------------------
    # 写回 Excel 表格的第 L 列
    # -------------------------
    col_l_index = 12  # L 列 = 第 12 列
    sheet.cell(row=1, column=col_l_index, value="数据集")  # 写表头
    for i, result in enumerate(results, start=2):
        sheet.cell(row=i, column=col_l_index, value=json.dumps(result, ensure_ascii=False))

    wb.save(excel_path)
    print(f"转换完成，结果已写入 {excel_path} 的 L 列（数据集）")

if __name__ == "__main__":
    input_excel_path = "C:\\Users\\autobio\\PycharmProjects\\study\\文件转换\\111.xlsx"
    generate_json_into_excel(input_excel_path)
