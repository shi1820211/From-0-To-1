import openpyxl
import json
import random
from collections import OrderedDict
import re


def parse_attributes(col_a, col_b, col_j=None, col_i=None):
    """
    解析 A 列内容，提取 sex, age, disease, physiologicalState 四类属性
    """
    pat_sex = None
    physiological_state = []
    disease = []
    age = {}

    has_sex, has_age, has_disease, has_state = 0, 0, 0, 0

    # ================== 先扫描 A 列 ==================
    if col_a:
        expressions = str(col_a).split(" and ")
        for expr in expressions:
            expr = expr.strip()

            # 1. sex
            if "sex" in expr:
                has_sex = 1
                match = re.search(r'sex\s*==\s*[\'"](\w+)[\'"]', expr)
                if match:
                    pat_sex = match.group(1)

            # 2. physiologicalState
            elif "in physiologicalState" in expr:
                has_state = 1
                match = re.search(r"[\'\"]([0-9a-fA-F-]+)[\'\"]\s+in physiologicalState", expr)
                if match:
                    physiological_state.append(match.group(1))

            # 3. disease
            elif "in disease" in expr:
                has_disease = 1
                match = re.search(r"[\'\"]([^\'\"]+)[\'\"]\s+in disease", expr)
                if match:
                    disease.append(match.group(1))

            # 4. age 范围 - 修复版：正确处理运算符
            elif "age" in expr:
                has_age = 1
                # 提取范围 (支持 <=, <, >=, >)
                match = re.findall(r"(\d+\.?\d*)?\s*([<>]=?)?\s*age\s*([<>]=?)?\s*(\d+\.?\d*)?", expr)
                if match:
                    for m in match:
                        low, op1, op2, high = m

                        # 处理运算符 - 修复逻辑
                        real_low = 0
                        real_high = 150

                        # 处理下界
                        if op1 in (">", ">=") or op2 in (">", ">="):
                            # 对于 age >= 14 的情况
                            if op2 in (">", ">=") and high:
                                real_low = float(high)  # 使用 high 作为下界
                            elif op1 in (">", ">=") and low:
                                real_low = float(low)  # 使用 low 作为下界

                        # 处理上界
                        if op1 in ("<", "<=") or op2 in ("<", "<="):
                            # 对于 age <= 30 的情况
                            if op2 in ("<", "<=") and high:
                                real_high = float(high)  # 使用 high 作为上界
                            elif op1 in ("<", "<=") and low:
                                real_high = float(low)  # 使用 low 作为上界

                        # 确保范围有效
                        if real_low > real_high:
                            real_low, real_high = real_high, real_low

                        # ========== 先取单位 ==========
                        unit = "year"
                        if col_b:
                            try:
                                # 使用 json.loads 替代 eval
                                if isinstance(col_b, str):
                                    unit_dict = json.loads(col_b.replace("'", "\""))
                                else:
                                    unit_dict = col_b
                                unit = unit_dict.get("age", "year")
                            except Exception:
                                pass

                        # ========== 按单位随机取值 ==========
                        if unit == "month":
                            val = round(random.uniform(real_low, real_high), 1)
                        elif unit == "day":
                            val = random.randint(int(real_low), int(real_high))
                        else:
                            val = random.randint(int(real_low), int(real_high))

                        if val <= 0:  # 必须大于 0
                            val = 1

                        age[unit] = val

    # ================== 如果 A 没有 age，走 J 列 ==================
    if not has_age and col_j:
        # 增强正则表达式以支持更多格式
        col_j_str = str(col_j)

        # 尝试匹配标准格式
        match = re.search(r"(\d+\.?\d*)?\s*<=?\s*age\s*<=?\s*(\d+\.?\d*)?", col_j_str)

        # 如果标准格式匹配失败，尝试匹配 "%s>=40" 格式
        if not match:
            match = re.search(r"%s\s*([<>]=?)\s*(\d+\.?\d*)", col_j_str)
            if match:
                operator = match.group(1)
                value = float(match.group(2)) if match.group(2) else 0

                # 根据运算符设置范围
                if operator in (">", ">="):
                    low = value
                    high = 150
                elif operator in ("<", "<="):
                    low = 0
                    high = value
                else:
                    low, high = 0, 150
            else:
                # fallback 默认 0-150
                low, high = 0, 150
        else:
            # 标准格式匹配成功
            low = float(match.group(1)) if match.group(1) else 0
            high = float(match.group(2)) if match.group(2) else 150

        unit = "year"
        if col_b:
            try:
                # 使用 json.loads 替代 eval
                if isinstance(col_b, str):
                    unit_dict = json.loads(col_b.replace("'", "\""))
                else:
                    unit_dict = col_b
                unit = unit_dict.get("age", "year")
            except Exception:
                pass

        if unit == "month":
            val = round(random.uniform(low, high), 1)
        elif unit == "day":
            val = random.randint(int(low), int(high))
        else:
            val = random.randint(int(low), int(high))

        if val <= 0:
            val = 1
        age[unit] = val

    # ================== sex fallback ==================
    if not has_sex:
        pat_sex = str(col_i).strip().lower() if col_i else "male"

    return pat_sex, age, physiological_state, disease


def generate_json_into_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        col_a, col_b, col_g, col_i, col_j = row[0], row[1], row[6], row[8], row[9]  # A,B,G,I,J 列

        # -------------------------
        # Step 1: 处理 items (G 列)
        # -------------------------
        items_dict = OrderedDict()
        if col_g:
            expressions = str(col_g).replace(" and ", "||").replace(" or ", "||").split("||")
            for expr in expressions:
                expr = expr.strip()
                if not expr:
                    continue
                parts = expr.split()
                if len(parts) >= 3:
                    uuid = parts[0].strip("()")
                    operator = parts[1].strip()
                    raw_value = parts[2].strip("\",(),\'")
                    value = raw_value

                    try:
                        num_val = float(raw_value)
                        if operator == ">":
                            value = num_val + 0.1
                        elif operator == "<":
                            value = num_val - 0.1
                        else:
                            value = num_val
                    except ValueError:
                        value = raw_value

                    items_dict[uuid] = {
                        "name": "",
                        "values": value,
                        "qualitativeResult": value,
                        "units": [],
                        "interval": "1-1"
                    }

        # -------------------------
        # Step 2: 解析属性 (A,B,J,I 列)
        # -------------------------
        pat_sex, age, physiological_state, disease = parse_attributes(col_a, col_b, col_j, col_i)

        # -------------------------
        # Step 3: 组合结果
        # -------------------------
        result_obj = {
            "input": json.dumps({
                "items": items_dict,
                "pat_sex": pat_sex,
                "age": age,
                "physiologicalState": physiological_state,
                "disease": disease
            }, ensure_ascii=False)
        }

        results.append(result_obj)

    # -------------------------
    # 写回 Excel 表格的第 L 列
    # -------------------------
    col_l_index = 12  # L 列
    sheet.cell(row=1, column=col_l_index, value="数据集")
    for i, result in enumerate(results, start=2):
        sheet.cell(row=i, column=col_l_index, value=json.dumps(result, ensure_ascii=False))

    wb.save(excel_path)
    print(f"转换完成，结果已写入 {excel_path} 的 L 列（数据集）")


if __name__ == "__main__":
    input_excel_path = "C:\\Users\\autobio\\PycharmProjects\\study\\文件转换\\112.xlsx"
    generate_json_into_excel(input_excel_path)