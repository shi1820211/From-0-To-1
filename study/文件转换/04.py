import openpyxl
import json
import random
from collections import OrderedDict
import re
import csv


#  读取单位映射（小写化索引）
def load_unit_map(csv_path):
    unit_dict = {}
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or len(row) < 2:
                continue
            unit_id, unit_name = row[0].strip().lower(), row[1].strip()  # 索引小写化
            if unit_id:
                unit_dict[unit_id] = unit_name
    return unit_dict


#  解析属性
def parse_attributes(col_a, col_b, col_j=None, col_i=None):
    pat_sex = None
    physiological_state = []
    disease = []
    age = {}

    has_sex, has_age, has_disease, has_state = 0, 0, 0, 0

    #  扫描 A 列
    if col_a:
        expressions = str(col_a).split(" and ")
        for expr in expressions:
            expr = expr.strip()

            # 1. sex
            if "sex" in expr:
                has_sex = 1
                match = re.search(r'sex\s*==\s*[\'"](\w+)[\'"]', expr)
                if match:
                    pat_sex = match.group(1)

            # 2. physiologicalState
            elif "in physiologicalState" in expr:
                has_state = 1
                match = re.search(r"[\'\"]([0-9a-fA-F-]+)[\'\"]\s+in physiologicalState", expr)
                if match:
                    physiological_state.append(match.group(1))

            # 3. disease
            elif "in disease" in expr:
                has_disease = 1
                match = re.search(r"[\'\"]([^\'\"]+)[\'\"]\s+in disease", expr)
                if match:
                    disease.append(match.group(1))

            # 4. age 范围
            elif "age" in expr:
                has_age = 1
                match = re.findall(r"(\d+\.?\d*)?\s*([<>]=?)?\s*age\s*([<>]=?)?\s*(\d+\.?\d*)?", expr)
                if match:
                    for m in match:
                        low, op1, op2, high = m
                        real_low, real_high = 0, 150

                        if op1 in (">", ">=") or op2 in (">", ">="):
                            if op2 in (">", ">=") and high:
                                real_low = float(high)
                            elif op1 in (">", ">=") and low:
                                real_low = float(low)
                        if op1 in ("<", "<=") or op2 in ("<", "<="):
                            if op2 in ("<", "<=") and high:
                                real_high = float(high)
                            elif op1 in ("<", "<=") and low:
                                real_high = float(low)

                        if real_low > real_high:
                            real_low, real_high = real_high, real_low

                        unit = "year"
                        if col_b:
                            try:
                                if isinstance(col_b, str):
                                    unit_dict = json.loads(col_b.replace("'", "\""))
                                else:
                                    unit_dict = col_b
                                unit = unit_dict.get("age", "year")
                            except Exception:
                                pass

                        if unit == "month":
                            val = round(random.uniform(real_low, real_high), 1)
                        elif unit == "day":
                            val = random.randint(int(real_low), int(real_high))
                        else:
                            val = random.randint(int(real_low), int(real_high))

                        if val <= 0:
                            val = 1
                        age[unit] = val

    # 如果 A 没有 age，走 J 列
    if not has_age and col_j:
        col_j_str = str(col_j)
        match = re.search(r"(\d+\.?\d*)?\s*<=?\s*%s\s*<=?\s*(\d+\.?\d*)?", col_j_str)
        if not match:
            match = re.search(r"%s\s*([<>]=?)\s*(\d+\.?\d*)", col_j_str)
            if match:
                operator = match.group(1)
                value = float(match.group(2)) if match.group(2) else 0
                if operator in (">", ">="):
                    low, high = value, 150
                elif operator in ("<", "<="):
                    low, high = 0, value
                else:
                    low, high = 0, 150
            else:
                low, high = 0, 150
        else:
            low = float(match.group(1)) if match.group(1) else 0
            high = float(match.group(2)) if match.group(2) else 150

        unit = "year"
        if col_b:
            try:
                if isinstance(col_b, str):
                    unit_dict = json.loads(col_b.replace("'", "\""))
                else:
                    unit_dict = col_b
                unit = unit_dict.get("age", "year")
            except Exception:
                pass

        if unit == "month":
            val = round(random.uniform(low, high), 1)
        elif unit == "day":
            val = random.randint(int(low), int(high))
        else:
            if int(high)-int(low)>1:
                val = random.randint(int(low), int(high))
            else :
                val =round(random.uniform(low, high), 1)

        if val <= 0:
            val = 1
        age[unit] = val

    if not has_sex:
        pat_sex = str(col_i).strip().lower() if col_i else "male"

    return pat_sex, age, physiological_state, disease


# 主逻辑
def generate_json_into_excel(excel_path, unit_csv_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # 读取单位字典（索引小写化）
    unit_dict = load_unit_map(unit_csv_path)

    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        col_a, col_b, col_e, col_g, col_i, col_j = row[0], row[1], row[4], row[6], row[8], row[9]

        # Step 1: 处理 items (G 列)
        items_dict = OrderedDict()
        if col_g:
            expressions = str(col_g).replace(" and ", "||").replace(" or ", "||").split("||")
            for expr in expressions:
                expr = expr.strip()
                if not expr:
                    continue
                parts = expr.split()
                if len(parts) >= 3:
                    uuid = parts[0].strip("()")
                    operator = parts[1].strip()
                    raw_value = parts[2].strip("\",(),\'")
                    value = raw_value
                    units_field = []

                    try:
                        # 数值型处理
                        num_val = float(raw_value)
                        if operator == ">":
                            value = num_val + 0.1
                        elif operator == "<":
                            value = num_val - 0.1
                        else:
                            value = num_val


                        # 查找单位
                        if col_e:
                            try:
                                if isinstance(col_e, str):
                                    unit_map = json.loads(col_e.replace("'", "\""))
                                else:
                                    unit_map = col_e
                            except Exception:
                                unit_map = {}
                        else:
                            unit_map = {}

                        unit_id = unit_map.get(uuid, "")
                        if unit_id:
                            unit_id_lower = unit_id.lower()
                            if unit_id_lower in unit_dict:
                                units_field = [{"unitName": unit_dict[unit_id_lower], "unitId": unit_id}]

                    except ValueError:
                        # 字符串型处理
                        if operator == "!=":
                            mapping = {
                                "降低": "升高",
                                "升高": "降低",
                                "阴性": "阳性",
                                "阳性": "阴性",
                                "增多": "减少",
                                "减少": "增多",
                                "正常": "不正常",
                            }
                            value = mapping.get(raw_value, raw_value)
                        else:
                            value = raw_value

                    items_dict[uuid] = {
                        "name": "",
                        "values": value,
                        "qualitativeResult": value,
                        "units": units_field,
                        "interval": "1-1"
                    }

        # Step 2: 解析属性
        pat_sex, age, physiological_state, disease = parse_attributes(col_a, col_b, col_j, col_i)

        # Step 3: 组合结果
        result_obj = {
            "input": json.dumps({
                "items": items_dict,
                "pat_sex": pat_sex,
                "age": age,
                "physiologicalState": physiological_state,
                "disease": disease
            }, ensure_ascii=False)
        }
        results.append(result_obj)

    # 写回 Excel L 列
    col_l_index = 12
    sheet.cell(row=1, column=col_l_index, value="数据集")
    for i, result in enumerate(results, start=2):
        sheet.cell(row=i, column=col_l_index, value=json.dumps(result, ensure_ascii=False))

    wb.save(excel_path)
    print(f"转换完成，结果已写入 {excel_path} 的 L 列（数据集）")


if __name__ == "__main__":
    input_excel_path = "C:\\Users\\autobio\\PycharmProjects\\study\\文件转换\\test_1.xlsx"
    unit_csv_path = "C:\\Users\\autobio\\PycharmProjects\\study\\文件转换\\单位.csv"
    generate_json_into_excel(input_excel_path, unit_csv_path)
