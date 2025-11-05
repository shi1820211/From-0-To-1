import openpyxl
from openpyxl.styles import PatternFill
import re

# 标黄样式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 提取区间条件
def parse_condition(text):
    # 提取第一个数字和符号
    pattern = r'([<>≤≥]+)\s*([0-9.]+)'
    matches = re.findall(pattern, text)
    intervals = []
    for op, num in matches:
        num = float(num)
        if op in ['<', '≤']:
            intervals.append(('upper', num, op))
        elif op in ['>', '≥']:
            intervals.append(('lower', num, op))
    return intervals

# 判断两个区间是否有交集（边界相等不算交集）
def intervals_intersect(int1, int2):
    lower1 = max([v for t, v, o in int1 if t=='lower'], default=float('-inf'))
    upper1 = min([v for t, v, o in int1 if t=='upper'], default=float('inf'))
    lower2 = max([v for t, v, o in int2 if t=='lower'], default=float('-inf'))
    upper2 = min([v for t, v, o in int2 if t=='upper'], default=float('inf'))
    return lower1 < upper2 and lower2 < upper1

# 判断数学逻辑错误
def has_logic_error(text):
    pattern = r'([0-9.]+)\s*<\s*.*\s*>\s*([0-9.]+)'
    return bool(re.search(pattern, text))

def process_excel(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 前四列交集列：I,L,O,R → 0起始为1，所以列号分别是 9, 12, 15, 18
    cols_intersect = [9, 12, 15, 18]

    # 后七列逻辑检查列：U,X,AA,AD,AG,AJ,AM → 列号 21,24,27,30,33,36,39
    cols_logic = [21, 24, 27, 30, 33, 36, 39]

    for row in ws.iter_rows(min_row=3):  # 假设数据从第三行开始
        # 处理交集
        intervals = []
        for col in cols_intersect:
            cell = row[col-1].value
            if cell:
                parsed = parse_condition(cell)
                intervals.append((col-1, parsed))

        # 四列互相比较
        for i in range(len(intervals)):
            for j in range(i+1, len(intervals)):
                if intervals[i][1] and intervals[j][1]:
                    if intervals_intersect(intervals[i][1], intervals[j][1]):
                        ws.cell(row=row[0].row, column=intervals[i][0]+1).fill = yellow_fill
                        ws.cell(row=row[0].row, column=intervals[j][0]+1).fill = yellow_fill

        # 处理数学逻辑错误
        for col in cols_logic:
            cell = row[col-1].value
            if cell and has_logic_error(cell):
                ws.cell(row=row[0].row, column=col).fill = yellow_fill

    wb.save(output_file)
    print(f"处理完成，新文件保存为 {output_file}")

# 使用示例
process_excel("表格核对-20250909.xlsx", "data_checked.xlsx")
