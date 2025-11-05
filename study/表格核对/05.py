import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 黄色填充
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 提取文本中的所有数字
def extract_numbers(text):
    numbers = re.findall(r"[-+]?\d*\.?\d+", text)
    return [float(num) for num in numbers]

# 解析区间，忽略文字
def parse_interval(text):
    if not text:
        return None
    text = text.replace("＜", "<").replace("＞", ">").replace("≤", "<=").replace("≥", ">=")
    text = text.replace(" ", "")

    nums = extract_numbers(text)

    # 双边区间 low < x < high
    if "<" in text and ">" not in text:
        if len(nums) >= 2:
            low, high = nums[0], nums[1]
            if low >= high:
                return None
            return (low, high)

    # 单边 <
    if "<" in text and len(nums) >= 1:
        return (float('-inf'), nums[0])

    # 单边 >
    if ">" in text and len(nums) >= 1:
        return (nums[0], float('inf'))

    return None

# 严格开区间交集判断，边界相等不算交集
def interval_overlap(a, b):
    if a is None or b is None:
        return False
    return max(a[0], b[0]) < min(a[1], b[1])

# 检查数学逻辑错误
def check_math_logic(text):
    if not text:
        return False
    text = text.replace("＜", "<").replace("＞", ">").replace(" ", "")
    nums = extract_numbers(text)
    # 检查出现 low<...>high 的情况
    match = re.search(r"([\d.]+)<.*?>\s*([\d.]+)", text)
    if match:
        low, high = float(match.group(1)), float(match.group(2))
        if low >= high:
            return True
    # 检查双边区间 low < x < high
    if "<" in text and len(nums) >= 2:
        if nums[0] >= nums[1]:
            return True
    return False

def process_excel(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active

    # I,L,O,R 列
    interval_cols = [9, 12, 15, 18]
    # U,X,AA,AD,AG,AJ,AM 列
    logic_cols = [21, 24, 27, 30, 33, 36, 39]

    for row in ws.iter_rows(min_row=3):
        # 解析四列区间
        intervals = []
        for col in interval_cols:
            cell = row[col-1]
            interval = parse_interval(str(cell.value))
            intervals.append(interval)

        # 四列相互比较交集
        for i in range(len(intervals)):
            for j in range(i+1, len(intervals)):
                if interval_overlap(intervals[i], intervals[j]):
                    row[interval_cols[i]-1].fill = yellow_fill
                    row[interval_cols[j]-1].fill = yellow_fill

        # 检查数学逻辑错误
        for col in logic_cols:
            cell = row[col-1]
            if cell.value and check_math_logic(str(cell.value)):
                cell.fill = yellow_fill

    wb.save(output_file)

if __name__ == "__main__":
    process_excel("表格核对-20250909.xlsx", "002-data_checked.xlsx")
