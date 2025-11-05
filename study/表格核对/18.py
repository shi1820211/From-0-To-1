import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill

# 💡 解析文本中的区间表达式
def parse_interval(s):
    """
    从字符串中提取区间。例如：
    - "<10" -> (-inf, 10)
    - ">35" -> (35, inf)
    - "low<浓度值<high" -> (low, high)
    """
    s = s.replace(" ", "")  # 去除空格

    # 格式1: <num
    less_match = re.match(r"<([0-9.]+)", s)
    if less_match:
        return (float('-inf'), float(less_match.group(1)))

    # 格式2: >num
    greater_match = re.match(r">([0-9.]+)", s)
    if greater_match:
        return (float(greater_match.group(1)), float('inf'))

    # 格式3: low<num<high
    range_match = re.match(r"([0-9.]+)<([^<=>]+)<([0-9.]+)", s)
    if range_match:
        low = float(range_match.group(1))
        high = float(range_match.group(3))
        return (low, high)

    return None  # 默认不识别

# 🔍 判断两个区间是否相交（开区间）
def is_intersecting(interval1, interval2):
    if interval1 is None or interval2 is None:
        return False
    left1, right1 = interval1
    left2, right2 = interval2
    return max(left1, left2) < min(right1, right2)

# 🤖 判断单元格的表达式是否存在错误逻辑（如 "low<浓度值>high"）
def has_logic_error(s):
    match = re.match(r"([0-9.]+)<([^<=>]+)>([0-9.]+)", s)
    if match:
        low = float(match.group(1))
        high = float(match.group(3))
        return low >= high  # 错误条件：low >= high
    return False

# 🎨 高亮单元格
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def highlight_cell(ws, row, col):
    ws.cell(row=row, column=col).fill = yellow_fill

# 📊 主处理函数
def process_excel(file_path, output_path):
    # 使用 pandas 读取数据
    df = pd.read_excel(file_path)

    # 使用 openpyxl 打开工作簿并获取对应工作表
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 定义需要检查的列 Excel 列索引（A=1, B=2, I=9, L=12 等等）
    cols_interval_check = [9 - 1, 12 - 1, 15 - 1, 18 - 1]     # Excel 列：I, L, O, R -> 8,11,14,17 (pandas 索引从0起)
    cols_logic_check = [21 - 1, 24 - 1, 27 - 1, 30 - 1, 33 - 1, 36 - 1, 39 - 1]  # U, X, AA, AD, AG, AJ, AM 列

    # 遍历每一行（跳过表头）
    for idx, row_data in df.iterrows():
        intervals = []
        # 提取区间
        for col in cols_interval_check:
            cell_val = df.iloc[idx, col]
            if isinstance(cell_val, str):
                interval = parse_interval(cell_val)
                intervals.append(interval)
            else:
                intervals.append(None)

        # 两两比较四列，判断是否有交集
        for i in range(len(intervals)):
            for j in range(i + 1, len(intervals)):
                if is_intersecting(intervals[i], intervals[j]):
                    # 标记交集的两个单元格为黄色
                    highlight_cell(ws, idx + 2, cols_interval_check[i] + 1)  # openpyxl 索引从 1 起始
                    highlight_cell(ws, idx + 2, cols_interval_check[j] + 1)

        # 数学逻辑错误检查（比如 low<浓度值>high）
        for col in cols_logic_check:
            cell_val = df.iloc[idx, col]
            if isinstance(cell_val, str) and has_logic_error(cell_val):
                highlight_cell(ws, idx + 2, col + 1)

    # 保存新文件
    wb.save(output_path)
    print(f"处理完成，新文件已保存至: {output_path}")

# 🚀 执行
if __name__ == "__main__":
    import sys
    import os

    input_file = '表格核对-20250909.xlsx'  # ← 您的原始文件名称
    output_file = '表格核对-处理后.xlsx'  # 导出文件名称

    if not os.path.exists(input_file):
        print(f"未找到文件: {input_file}, 请放在当前路径下")
    else:
        process_excel(input_file, output_file)
