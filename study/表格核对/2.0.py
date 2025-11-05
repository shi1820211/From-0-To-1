import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 提取单元格中最后一个数学条件，只关注符号+数字
def extract_condition(cell_value):
    if not isinstance(cell_value, str):
        return None
    # 匹配全角/半角符号 + 数字
    matches = re.findall(r'[<>≤≥＜＞]\s*\d+\.?\d*', cell_value)
    if not matches:
        return None
    cond = matches[-1]
    # 统一符号
    cond = cond.replace('＜', '<').replace('＞', '>') \
               .replace('≤', '<=').replace('≥', '>=')
    # 转为区间
    try:
        if '<=' in cond:
            val = float(cond.split('<=')[-1])
            return (-float('inf'), val)
        elif '>=' in cond:
            val = float(cond.split('>=')[-1])
            return (val, float('inf'))
        elif '<' in cond:
            val = float(cond.split('<')[-1])
            return (-float('inf'), val)
        elif '>' in cond:
            val = float(cond.split('>')[-1])
            return (val, float('inf'))
        else:
            return None
    except:
        return None

# 判断区间是否有交集
def has_intersection(c1, c2):
    if not c1 or not c2:
        return False
    return not (c1[1] < c2[0] or c2[1] < c1[0])

# 检查数学逻辑错误
def has_math_error(cell_value):
    if not isinstance(cell_value, str):
        return False
    # 找到最后一个条件
    cond = extract_condition(cell_value)
    if not cond:
        return False
    # 如果出现 "<" 和 ">" 混用的情况，标黄
    # 例如: 25<浓度值>30
    if re.search(r'\d+\s*<.*>\s*\d+', cell_value):
        return True
    return False

def process_excel(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active

    header_row = 3
    headers = [cell.value for cell in ws[header_row]]
    zhao_cols = [i+1 for i, h in enumerate(headers) if h == "赵"]

    # 前5列处理：信息2~信息5
    for row in range(header_row+1, ws.max_row+1):
        conds = []
        for idx in zhao_cols[1:5]:
            val = ws.cell(row=row, column=idx).value
            conds.append((idx, extract_condition(val)))
        # 四列相互比较
        for i in range(len(conds)):
            for j in range(i+1, len(conds)):
                idx1, c1 = conds[i]
                idx2, c2 = conds[j]
                if has_intersection(c1, c2):
                    ws.cell(row=row, column=idx1).fill = yellow_fill
                    ws.cell(row=row, column=idx2).fill = yellow_fill

    # 后7列检查逻辑错误：信息6~信息12
    for row in range(header_row+1, ws.max_row+1):
        for idx in zhao_cols[5:]:
            val = ws.cell(row=row, column=idx).value
            if has_math_error(val):
                ws.cell(row=row, column=idx).fill = yellow_fill

    wb.save(output_file)
    print(f"处理完成，已生成新文件: {output_file}")


# 使用示例
process_excel("表格核对-20250909.xlsx", "data_checked.xlsx")
