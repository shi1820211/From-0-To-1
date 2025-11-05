import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 黄色填充
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# 提取文本中第一个数字，用作区间边界
def extract_first_number(text):
    match = re.search(r"[\d.]+", text)
    return float(match.group()) if match else None


# 解析区间，支持单边和双边
def parse_interval(text):
    if not text:
        return None
    text = text.replace("＜", "<").replace("＞", ">").replace("≤", "<=").replace("≥", ">=")
    text = text.replace(" ", "")

    # 双边区间 low<x<high
    match2 = re.findall(r"([\d.]+)<.*<([\d.]+)", text)
    if match2:
        low, high = float(match2[0][0]), float(match2[0][1])
        if low >= high:
            return None
        return (low, high)

    # 单边 <
    if "<" in text:
        num = extract_first_number(text)
        if num is not None:
            return (float('-inf'), num)

    # 单边 >
    if ">" in text:
        num = extract_first_number(text)
        if num is not None:
            return (num, float('inf'))

    return None


# 严格开区间交集判断，边界相等不算交集
def interval_overlap(a, b):
    if a is None or b is None:
        return False
    return max(a[0], b[0]) < min(a[1], b[1])


# 检查数学逻辑错误，只针对双边区间
def check_math_logic(text):
    if not text:
        return False
    text = text.replace("＜", "<").replace("＞", ">").replace(" ", "")
    match = re.findall(r"([\d.]+)<.*<([\d.]+)", text)
    for m in match:
        low, high = float(m[0]), float(m[1])
        if low >= high:
            return True
    return False


def process_excel(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active

    # I,L,O,R 四列（区间交集判断）
    interval_cols = [9, 12, 15, 18]
    # U,X,AA,AD,AG,AJ,AM 七列（数学逻辑检查）
    logic_cols = [21, 24, 27, 30, 33, 36, 39]

    for row in ws.iter_rows(min_row=3):
        # 解析四列区间
        intervals = []
        for col in interval_cols:
            cell = row[col - 1]
            interval = parse_interval(str(cell.value))
            intervals.append(interval)

        # 四列相互比较交集
        for i in range(len(intervals)):
            for j in range(i + 1, len(intervals)):
                if interval_overlap(intervals[i], intervals[j]):
                    row[interval_cols[i] - 1].fill = yellow_fill
                    row[interval_cols[j] - 1].fill = yellow_fill

        # 检查数学逻辑错误（双边区间）
        for col in logic_cols:
            cell = row[col - 1]
            if cell.value and check_math_logic(str(cell.value)):
                cell.fill = yellow_fill

    wb.save(output_file)


if __name__ == "__main__":
    process_excel("表格核对-20250909.xlsx", "005-data_checked.xlsx")
