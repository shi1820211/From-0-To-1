import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 黄色填充
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 提取单区间的函数
def extract_interval(text):
    """
    返回开区间(low, high)的元组
    - '<num' -> (-inf, num)
    - '>num' -> (num, inf)
    - 'low<num<high' -> (low, high)
    如果无法解析返回 None
    """
    # 查找所有数字（整数或浮点数）
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
    if '<' in text and '>' not in text:  # <num
        return (-float('inf'), nums[-1])
    elif '>' in text and '<' not in text:  # >num
        return (nums[0], float('inf'))
    elif '<' in text and '>' in text:  # low<num<high
        if len(nums) >= 2:
            return (nums[0], nums[1])
    return None

# 判断两个开区间是否有交集
def intervals_intersect(int1, int2):
    if int1 is None or int2 is None:
        return False
    return max(int1[0], int2[0]) < min(int1[1], int2[1])  # 严格开区间

# 检查数学逻辑错误，例如 25<浓度值>30
def check_math_error(text):
    # 匹配 low<...>high 的情况
    pattern = re.compile(r"\d+\s*<[^<>\n]*>\s*\d+")
    return bool(pattern.search(text))

# 主函数
def process_excel(file_path, output_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # I L O R 列索引（1-based）
    interval_cols = [9, 12, 15, 18]

    # 后七列 U X AA AD AG AJ AM（1-based）
    math_cols = [21, 24, 27, 30, 33, 36, 39]

    for row in ws.iter_rows(min_row=3):  # 数据从第3行开始
        # ---------------- 规则1：区间交集 ----------------
        intervals = []
        for col in interval_cols:
            text = str(row[col-1].value)
            interval = extract_interval(text)
            intervals.append(interval)

        # 四列两两比较是否有交集
        n = len(intervals)
        for i in range(n):
            for j in range(i+1, n):
                if intervals_intersect(intervals[i], intervals[j]):
                    # 标黄
                    row[interval_cols[i]-1].fill = yellow_fill
                    row[interval_cols[j]-1].fill = yellow_fill

        # ---------------- 规则2：数学逻辑 ----------------
        for col in math_cols:
            text = str(row[col-1].value)
            if check_math_error(text):
                row[col-1].fill = yellow_fill

    wb.save(output_path)

# 调用
process_excel("表格核对-20250909.xlsx", "111data_checked.xlsx")

