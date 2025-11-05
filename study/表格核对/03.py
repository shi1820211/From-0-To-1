import openpyxl
from openpyxl.styles import PatternFill
import re

# 标黄样式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 中文和英文符号映射
symbol_map = {'<': '<', '＞': '>', '>': '>', '＜': '<', '≤': '<=', '≥': '>='}


# 解析单元格文本为区间 (lower, upper)
def parse_condition(text):
    if not text or str(text).strip() == "":
        return None
    text = str(text).replace(" ", "")

    # 双边区间，例如 0.5<浓度值<1.5 或 4＜样本结果＜10
    pattern = r'([0-9.]+)\s*[<＞≤≥]\s*.*?\s*[<＞≤≥]\s*([0-9.]+)'
    match = re.search(pattern, text)
    if match:
        lower = float(match.group(1))
        upper = float(match.group(2))
        return (lower, upper)

    # 单边区间，例如 <1 或 >0.5
    pattern_single = r'([<＞≤≥])\s*([0-9.]+)'
    match_single = re.search(pattern_single, text)
    if match_single:
        op, num = match_single.groups()
        num = float(num)
        op_std = symbol_map.get(op, op)
        if op_std in ['<', '<=']:
            return (float('-inf'), num)
        elif op_std in ['>', '>=']:
            return (num, float('inf'))

    return None


# 判断两个区间是否有交集（边界相等不算交集）
def intervals_intersect(int1, int2):
    if not int1 or not int2:
        return False
    return int1[0] < int2[1] and int2[0] < int1[1]


# 检查数学逻辑错误
def has_logic_error(text):
    if not text or str(text).strip() == "":
        return False
    # 检查类似 "25<浓度值>30" 或 "4＜样本结果＜2"
    pattern = r'([0-9.]+)\s*[<＞≤≥]\s*.*?\s*[<＞≤≥]\s*([0-9.]+)'
    matches = re.findall(pattern, str(text))
    for lower, upper in matches:
        if float(lower) >= float(upper):
            return True
    return False


def process_excel(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # I,L,O,R 四列互相比较 (列号从1开始)
    cols_intersect = [9, 12, 15, 18]
    # U,X,AA,AD,AG,AJ,AM 七列逻辑检查
    cols_logic = [21, 24, 27, 30, 33, 36, 39]

    for row in ws.iter_rows(min_row=3):  # 假设数据从第三行开始
        # 解析四列区间
        intervals = []
        for col in cols_intersect:
            cell_value = row[col - 1].value
            interval = parse_condition(str(cell_value)) if cell_value else None
            intervals.append((col - 1, interval))

        # 两两比较区间交集
        for i in range(len(intervals)):
            for j in range(i + 1, len(intervals)):
                col_i, int_i = intervals[i]
                col_j, int_j = intervals[j]
                if int_i and int_j and intervals_intersect(int_i, int_j):
                    ws.cell(row=row[0].row, column=col_i + 1).fill = yellow_fill
                    ws.cell(row=row[0].row, column=col_j + 1).fill = yellow_fill

        # 后七列数学逻辑错误检查
        for col in cols_logic:
            cell_value = row[col - 1].value
            if cell_value and has_logic_error(str(cell_value)):
                ws.cell(row=row[0].row, column=col).fill = yellow_fill

    wb.save(output_file)
    print(f"处理完成，新文件保存为 {output_file}")


# 使用示例
process_excel("表格核对-20250909.xlsx", "01-data_checked.xlsx")
