import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 提取条件区间
def extract_conditions(text):
    if not text or text in ["#N/A", "1"]:
        return None

    # 提取不等式部分
    matches = re.findall(r"([<>]=?|=)?\s*(-?\d+\.?\d*)", text)
    if not matches:
        return None

    # 转换成范围
    nums = []
    for op, val in matches:
        try:
            nums.append((op, float(val)))
        except:
            continue

    if not nums:
        return None

    # 单个条件
    if len(nums) == 1:
        op, val = nums[0]
        if op in ["<", "≤"]:
            return (-float("inf"), val)
        elif op in [">", "≥"]:
            return (val, float("inf"))
        else:
            return (val, val)

    # 双条件  (区间)
    elif len(nums) == 2:
        (op1, v1), (op2, v2) = nums
        return (min(v1, v2), max(v1, v2))

    return None

# 判断区间交集
def has_intersection(cond1, cond2):
    if not cond1 or not cond2:
        return False
    return not (cond1[1] < cond2[0] or cond2[1] < cond1[0])

# 检查数学逻辑错误
def has_math_error(text):
    if not text or text in ["#N/A", "1"]:
        return False
    # 错误模式: a < x > b 或 a > x < b
    return bool(re.search(r"\d+\s*[<>]=?\s*[^<>=\d]+\s*[<>]=?\s*\d+", text) and
                re.search(r"<[^<]*>", text))

def process_excel(input_file, output_file):
    wb = load_workbook('表格核对-20250909.xlsx')
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 找到“赵”列的索引
    header_row = 3  # 第3行为表头
    headers = [cell.value for cell in ws[header_row]]
    zhao_cols = [i+1 for i, h in enumerate(headers) if h == "赵"]

    if len(zhao_cols) < 12:
        print("警告: 找到的‘赵’列少于12列")
    else:
        # 前5列处理 (信息2~信息5 = zhao_cols[1:5])
        for row in range(header_row+1, ws.max_row+1):
            conds = []
            for idx in zhao_cols[1:5]:  # 信息2-信息5
                val = ws.cell(row=row, column=idx).value
                conds.append((idx, extract_conditions(str(val) if val else None)))

            # 四列相互比较
            for i in range(len(conds)):
                for j in range(i+1, len(conds)):
                    idx1, c1 = conds[i]
                    idx2, c2 = conds[j]
                    if has_intersection(c1, c2):
                        ws.cell(row=row, column=idx1).fill = yellow_fill
                        ws.cell(row=row, column=idx2).fill = yellow_fill

        # 后7列检查逻辑错误 (信息6~信息12)
        for row in range(header_row+1, ws.max_row+1):
            for idx in zhao_cols[5:]:
                val = ws.cell(row=row, column=idx).value
                if val and has_math_error(str(val)):
                    ws.cell(row=row, column=idx).fill = yellow_fill

    wb.save(output_file)
    print(f"处理完成，已生成新文件: {output_file}")


# 使用示例
process_excel("data.xlsx", "data_checked.xlsx")
